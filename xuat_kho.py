import streamlit as st
import pandas as pd
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# --- CẤU HÌNH TRANG WEB ---
st.set_page_config(page_title="Xuất Phiếu Kho", layout="wide")

st.title("🖨️ Phần Mềm Xuất Phiếu Kho Online")

# --- KHỞI TẠO SESSION STATE (Để lưu danh sách hàng khi trang web reload) ---
if 'cart' not in st.session_state:
    st.session_state.cart = []

# --- PHẦN 1: THÔNG TIN CHUNG ---
with st.container():
    st.subheader("1. Thông tin chung")
    col1, col2 = st.columns(2)
    
    with col1:
        khach_hang = st.text_input("Tên khách hàng", placeholder="VD: Quầy thuốc Hoa Lan")
        dia_chi = st.text_input("Địa chỉ", placeholder="VD: IaKrai")
        sdt = st.text_input("Số điện thoại")
    
    with col2:
        nv_ban = st.text_input("Nhân viên bán hàng", value="Phan Ngọc Cường")
        ngay_xuat = st.date_input("Ngày xuất", datetime.now())

# --- PHẦN 2: NHẬP HÀNG HÓA ---
st.subheader("2. Nhập chi tiết hàng hóa")

# Tạo form để khi nhấn Enter hoặc nút Thêm thì mới xử lý
with st.form("entry_form", clear_on_submit=True):
    c1, c2, c3, c4, c5 = st.columns([3, 1, 1, 1.5, 2])
    
    with c1:
        ten_hang = st.text_input("Tên hàng")
    with c2:
        dvt = st.text_input("ĐVT", value="Hộp")
    with c3:
        sl = st.number_input("Số lượng", min_value=1, value=1, step=1)
    with c4:
        don_gia = st.number_input("Đơn giá", min_value=0, value=0, step=1000)
    with c5:
        ghi_chu = st.text_input("Ghi chú")

    submit = st.form_submit_button("➕ Thêm vào danh sách")

    if submit:
        if ten_hang:
            thanh_tien = sl * don_gia
            st.session_state.cart.append({
                "STT": len(st.session_state.cart) + 1,
                "Tên Hàng": ten_hang,
                "ĐVT": dvt,
                "Số Lượng": sl,
                "Đơn Giá": don_gia,
                "Thành Tiền": thanh_tien,
                "Ghi Chú": ghi_chu
            })
            st.success(f"Đã thêm: {ten_hang}")
        else:
            st.error("Vui lòng nhập tên hàng!")

# --- PHẦN 3: DANH SÁCH & XUẤT FILE ---
st.divider()
st.subheader("3. Xem trước phiếu & Xuất Excel")

if len(st.session_state.cart) > 0:
    # Hiển thị bảng dữ liệu
    df = pd.DataFrame(st.session_state.cart)
    
    # Định dạng hiển thị số tiền cho đẹp trên Web
    df_display = df.copy()
    df_display['Đơn Giá'] = df_display['Đơn Giá'].apply(lambda x: "{:,.0f}".format(x))
    df_display['Thành Tiền'] = df_display['Thành Tiền'].apply(lambda x: "{:,.0f}".format(x))
    
    st.dataframe(df_display, use_container_width=True)

    # Tính tổng tiền
    tong_tien = sum(item['Thành Tiền'] for item in st.session_state.cart)
    st.markdown(f"<h3 style='text-align: right; color: red;'>Tổng cộng: {tong_tien:,.0f} VNĐ</h3>", unsafe_allow_html=True)

    # Nút xóa danh sách
    if st.button("🗑️ Xóa toàn bộ danh sách"):
        st.session_state.cart = []
        st.rerun()

    # --- HÀM TẠO FILE EXCEL (LOGIC CŨ) ---
    def generate_excel():
        wb = Workbook()
        ws = wb.active
        ws.title = "PhieuXuatKho"

        # Styles
        font_bold = Font(name='Times New Roman', bold=True, size=11)
        font_normal = Font(name='Times New Roman', size=11)
        font_title = Font(name='Times New Roman', bold=True, size=16)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Header info
        ws.merge_cells('A2:G2')
        ws['A2'] = "PHIẾU XUẤT KHO BÁN HÀNG"
        ws['A2'].font = font_title
        ws['A2'].alignment = Alignment(horizontal='center')

        formatted_date = ngay_xuat.strftime("%d/%m/%Y")
        ws.merge_cells('A3:G3')
        ws['A3'] = f"Ngày: {formatted_date}"
        ws['A3'].font = font_normal
        ws['A3'].alignment = Alignment(horizontal='center')

        ws['A4'] = f"Tên khách hàng: {khach_hang}"; ws['A4'].font = font_normal
        ws['A5'] = f"Địa chỉ: {dia_chi}"; ws['A5'].font = font_normal
        ws['A6'] = f"SĐT: {sdt}"; ws['A6'].font = font_normal
        ws['A7'] = f"NV bán hàng: {nv_ban}"; ws['A7'].font = font_normal

        # Table Header
        headers = ["STT", "TÊN HÀNG", "Đ.V TÍNH", "SỐ LƯỢNG", "ĐƠN GIÁ", "THÀNH TIỀN", "GHI CHÚ"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col, value=header)
            cell.font = font_bold; cell.border = thin_border; cell.alignment = Alignment(horizontal='center')

        # Table Data
        curr_row = 10
        for item in st.session_state.cart:
            row_vals = [item["STT"], item["Tên Hàng"], item["ĐVT"], item["Số Lượng"], item["Đơn Giá"], item["Thành Tiền"], item["Ghi Chú"]]
            for col, val in enumerate(row_vals, 1):
                cell = ws.cell(row=curr_row, column=col, value=val)
                cell.font = font_normal; cell.border = thin_border
                if col in [4, 5, 6]: # Số tiền
                    cell.number_format = '#,##0'
            curr_row += 1

        # Total row
        ws.merge_cells(f'A{curr_row}:E{curr_row}')
        ws[f'A{curr_row}'] = "CỘNG:"
        ws[f'A{curr_row}'].font = font_bold; ws[f'A{curr_row}'].border = thin_border
        for i in range(1, 6): ws.cell(row=curr_row, column=i).border = thin_border
        
        ws.cell(row=curr_row, column=6, value=tong_tien).number_format = '#,##0'
        ws.cell(row=curr_row, column=6).font = font_bold; ws.cell(row=curr_row, column=6).border = thin_border
        ws.cell(row=curr_row, column=7).border = thin_border

        # Signatures
        f_row = curr_row + 2
        sigs = ["Người mua hàng", "Thủ kho", "Kế toán"]
        positions = ['A', 'C', 'E']
        for i, sig in enumerate(sigs):
            ws.merge_cells(f'{positions[i]}{f_row}:{chr(ord(positions[i])+1)}{f_row}')
            c = ws[f'{positions[i]}{f_row}']
            c.value = sig; c.font = font_normal; c.alignment = Alignment(horizontal='center')
            
            # (ký họ tên)
            ws.merge_cells(f'{positions[i]}{f_row+1}:{chr(ord(positions[i])+1)}{f_row+1}')
            c2 = ws[f'{positions[i]}{f_row+1}']
            c2.value = "(ký, họ tên)"; c2.font = Font(name='Times New Roman', italic=True, size=10); c2.alignment = Alignment(horizontal='center')

        # Columns width
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # --- NÚT TẢI FILE ---
    file_name = f"Phieu_{khach_hang}_{datetime.now().strftime('%d%m%Y')}.xlsx"
    st.download_button(
        label="📥 TẢI XUỐNG FILE EXCEL",
        data=generate_excel(),
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary" # Nút màu đỏ nổi bật
    )

else:
    st.info("Chưa có hàng hóa nào trong danh sách. Hãy nhập ở trên!")